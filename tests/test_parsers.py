"""Tests for Excel and PDF parsers."""

import json
import os
import sys
import tempfile
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).parent.parent))

from tools.parsers.excel_parser import ExcelParserTool


def make_sample_excel(path: str):
    """Create a minimal Excel file with DRE data for testing."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DRE"

    # Header row with years
    ws["A1"] = "Conta"
    ws["B1"] = "2022"
    ws["C1"] = "2023"
    ws["D1"] = "2024"

    # DRE rows (Brazilian format)
    rows = [
        ("Receita Líquida", "80.000.000", "96.000.000", "116.000.000"),
        ("Custo dos Produtos Vendidos", "(48.000.000)", "(56.000.000)", "(67.000.000)"),
        ("Lucro Bruto", "32.000.000", "40.000.000", "49.000.000"),
        ("EBITDA", "18.000.000", "23.500.000", "29.000.000"),
        ("Depreciação e Amortização", "(4.000.000)", "(4.800.000)", "(5.800.000)"),
        ("Lucro Líquido", "7.260.000", "10.164.000", "12.936.000"),
    ]
    for r_idx, (label, *vals) in enumerate(rows, 2):
        ws.cell(row=r_idx, column=1, value=label)
        for c_idx, val in enumerate(vals, 2):
            ws.cell(row=r_idx, column=c_idx, value=val)

    wb.save(path)


class TestExcelParser:
    def test_parse_dre_basic(self, tmp_path):
        excel_path = str(tmp_path / "test_dre.xlsx")
        make_sample_excel(excel_path)

        tool = ExcelParserTool()
        result_str = tool._run(file_path=excel_path)
        result = json.loads(result_str)

        assert "entities" in result
        assert not result.get("error"), f"Error: {result.get('error')}"

    def test_parse_nonexistent_file(self):
        tool = ExcelParserTool()
        result = json.loads(tool._run(file_path="/nonexistent/file.xlsx"))
        assert "error" in result

    def test_to_float_conversions(self):
        tool = ExcelParserTool()
        assert tool._to_float("1.234.567,89") == 1_234_567.89
        assert tool._to_float("(500.000,00)") == -500_000.0
        assert tool._to_float("R$ 1.000,00") == 1_000.0
        assert tool._to_float("abc") == 0.0
        assert tool._to_float("") == 0.0

    def test_is_numeric(self):
        tool = ExcelParserTool()
        assert tool._is_numeric("1.234.567,89")
        assert tool._is_numeric("(500.000)")
        assert not tool._is_numeric("Receita Líquida")
        assert not tool._is_numeric("")

    def test_detect_years_in_header(self, tmp_path):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Conta"
        ws["B1"] = "2022"
        ws["C1"] = "2023"
        ws["A2"] = "Receita"
        ws["B2"] = "100"
        ws["C2"] = "120"
        path = str(tmp_path / "years.xlsx")
        wb.save(path)

        tool = ExcelParserTool()
        import pandas as pd
        df = pd.read_excel(path, header=None, dtype=str).fillna("")
        years = tool._extract_years(df)
        assert "2022" in years
        assert "2023" in years
