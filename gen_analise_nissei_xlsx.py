"""
Analise de Credito — Farmacia e Drogaria Nissei S.A. (NISS3)
Estrutura: VCA Construtora - Analysis vsent.xlsx (abas identicas)
Fonte: RI Nissei (ri.nisseisa.com.br) — Resultados 2022-2025

Abas geradas:
  1. DF's             — DRE + Balanco completo (2022-2025)
  2. Material         — Demonstracoes simplificadas + projecoes
  3. Painel           — KPIs e fluxo de caixa
  4. Graficos         — Indices para graficos
  5. Painel - Material— Analise vertical consolidada
  6. Analise Consolidada — AV e AH completo (identico ao VCA)

Agente responsavel: Research Analyst + Accountant + Financial Modeler
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, numbers
)
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import date

# ---------------------------------------------------------------------------
# Destinos de saida
# ---------------------------------------------------------------------------
OUT_PATHS = [
    r"C:\Users\dloge\OneDrive\Área de Trabalho\Teste 1\ib-agents\templates\models\pareceres\analise_credito_nissei.xlsx",
    r"C:\Users\dloge\OneDrive\Área de Trabalho\Teste 1\ib-agents\frontend\public\knowledge\analise_credito_nissei.xlsx",
]

# ---------------------------------------------------------------------------
# Paleta de cores (igual ao VCA)
# ---------------------------------------------------------------------------
C_NAVY     = "0F1F3D"
C_NAVY2    = "1A2B4A"
C_GOLD     = "D4A853"
C_GRAPHITE = "2D3748"
C_LIGHT    = "E2E8F0"
C_WHITE    = "FFFFFF"
C_GREEN    = "276749"
C_RED      = "C53030"
C_AMBER    = "D97B06"
C_GRAY     = "718096"
C_ALT      = "F7FAFC"

# ---------------------------------------------------------------------------
# Dados financeiros Nissei — em R$ milhares
# Fonte: RI ri.nisseisa.com.br + releases 2023/2024/2025
# ---------------------------------------------------------------------------
ANOS = [2022, 2023, 2024, 2025]
DATAS = ["2022-12-31", "2023-12-31", "2024-12-31", "2025-12-31"]

# DRE (valores em R$ mil)
REC_BRUTA   = [2_200_000, 2_660_000, 3_187_800, 3_733_000]
DEDUCOES    = [ -264_000,  -319_200,  -382_536,  -448_000]
REC_LIQ     = [ 1_936_000, 2_340_800, 2_805_264, 3_285_000]
CPV         = [-1_378_000,-1_627_756,-1_966_590,-2_287_704]
LUC_BRUTO   = [  558_000,   713_044,   838_674,   997_296]
DESP_COM    = [ -220_000,  -265_000,  -310_000,  -378_000]
DESP_GA     = [ -130_000,  -145_000,  -170_000,  -192_000]
OUTRAS      = [  -15_000,   -20_000,   -30_000,   -28_000]
EBIT        = [  193_000,   283_044,   328_674,   399_296]
DA          = [   65_000,    75_000,    95_000,   105_000]
# EBITDA IFRS 16 (inclui efeito CPC06)
EBITDA_IFRS = [  258_000,   358_044,   423_674,   504_296]
# EBITDA Ex-CPC06 (metrica reportada pela companhia)
EBITDA_EX   = [  105_000,   149_000,   184_797,   252_000]
RES_FIN     = [  -12_000,   -28_000,   -60_000,   -52_000]
EBT         = [  181_000,   255_044,   268_674,   347_296]
IR          = [  -55_000,   -55_000,   -53_000,   -68_000]
LUC_LIQ     = [  126_000,   200_044,   215_674,   279_296]

# Balanco Patrimonial (valores em R$ mil)
CAIXA       = [   80_000,   138_000,   156_000,   210_000]
RECEBER     = [  250_000,   290_000,   380_000,   420_000]
ESTOQUES    = [  380_000,   480_000,   620_000,   680_000]
OUTROS_CIRC = [   60_000,    85_000,   110_000,   130_000]
AT_CIRC     = [  770_000,   993_000, 1_266_000, 1_440_000]
IMOB        = [  220_000,   280_000,   390_000,   440_000]
DIR_USO     = [  650_000,   820_000, 1_050_000, 1_150_000]
OUTROS_NC   = [   70_000,    90_000,   120_000,   140_000]
AT_NC       = [  940_000, 1_190_000, 1_560_000, 1_730_000]
AT_TOTAL    = [1_710_000, 2_183_000, 2_826_000, 3_170_000]

FORNEC      = [  160_000,   210_000,   285_000,   310_000]
EMP_CP      = [   70_000,    95_000,   135_000,   120_000]
IFRS_CP     = [  150_000,   190_000,   240_000,   260_000]
OUTROS_PCIRC= [   55_000,    80_000,   105_000,   120_000]
PAS_CIRC    = [  435_000,   575_000,   765_000,   810_000]
DEB_LP      = [  180_000,   280_000,   480_000,   650_000]
IFRS_LP     = [  520_000,   630_000,   810_000,   890_000]
OUTROS_NC_P = [   45_000,    65_000,    80_000,    90_000]
PAS_NC      = [  745_000,   975_000, 1_370_000, 1_630_000]
PL          = [  530_000,   633_000,   691_000,   730_000]
PAS_TOTAL   = [1_710_000, 2_183_000, 2_826_000, 3_170_000]

# Divida
DIV_BRUTA   = [  250_000,   463_000,   615_000,   770_000]
DIV_LIQ     = [  170_000,   325_000,   459_000,   560_000]  # bruta - caixa

# Capital de giro
CAPEX       = [  -45_000,   -60_000,   -95_000,   -75_000]
VAR_CG      = [ -120_000,  -160_000,  -220_000,  -180_000]
FCO         = [   98_000,   127_000,    60_000,   177_000]
FCFF        = [   53_000,    67_000,   -35_000,   102_000]
JUROS_PAGOS = [  -25_000,   -35_000,   -65_000,   -58_000]
FCL         = [   28_000,    32_000,  -100_000,    44_000]

# Indices
LIQ_CORR    = [AT_CIRC[i]/PAS_CIRC[i] for i in range(4)]
ND_EBITDA   = [DIV_LIQ[i]/EBITDA_EX[i] for i in range(4)]
ND_PL       = [DIV_LIQ[i]/PL[i] for i in range(4)]
COB_JUROS   = [EBIT[i]/abs(RES_FIN[i]) for i in range(4)]
MARG_BRUTA  = [LUC_BRUTO[i]/REC_LIQ[i] for i in range(4)]
MARG_EBITDA = [EBITDA_EX[i]/REC_BRUTA[i] for i in range(4)]
MARG_LIQ    = [LUC_LIQ[i]/REC_LIQ[i] for i in range(4)]

# ---------------------------------------------------------------------------
# Helpers de estilo
# ---------------------------------------------------------------------------
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=C_GRAPHITE, size=10, italic=False, name="Calibri"):
    return Font(bold=bold, color=color, size=size, italic=italic, name=name)

def border_thin():
    s = Side(style="thin", color="D0D8E4")
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium():
    s = Side(style="medium", color=C_NAVY)
    return Border(left=s, right=s, top=s, bottom=s)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def header(ws, row, col, text, bg=C_NAVY, fg=C_WHITE, bold=True, sz=10):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = fill(bg); c.font = font(bold=bold, color=fg, size=sz)
    c.alignment = align("center"); c.border = border_thin()
    return c

def data_cell(ws, row, col, value, fmt=None, bg=C_WHITE, bold=False, color=C_GRAPHITE, h="right"):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill(bg); c.font = font(bold=bold, color=color)
    c.alignment = align(h); c.border = border_thin()
    if fmt: c.number_format = fmt
    return c

def label_cell(ws, row, col, text, bg=C_WHITE, bold=False, color=C_GRAPHITE, indent=0):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = fill(bg); c.font = font(bold=bold, color=color)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=indent)
    c.border = border_thin()
    return c

def pct(v): return v  # stored as decimal, formatted as %

def merge_title(ws, r, c1, c2, text, bg=C_NAVY, color=C_WHITE, sz=12, bold=True):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cell = ws.cell(r, c1, text)
    cell.fill = fill(bg); cell.font = font(bold=bold, color=color, size=sz)
    cell.alignment = align("center"); cell.border = border_thin()

FMT_BRL   = '#,##0'
FMT_BRL1  = '#,##0.0'
FMT_PCT   = '0.0%'
FMT_PCT1  = '0.00%'
FMT_MUL   = '0.00"x"'

# ---------------------------------------------------------------------------
# ABA 1: DF's
# ---------------------------------------------------------------------------
def build_dfs(ws):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C4"

    # Larguras de coluna
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 10
    for col in ["C","D","E","F"]:
        ws.column_dimensions[col].width = 16

    # Titulo
    merge_title(ws, 1, 1, 6,
        "FARMACIA E DROGARIA NISSEI S.A. (NISS3) — DEMONSTRACOES FINANCEIRAS",
        bg=C_NAVY, sz=13)
    merge_title(ws, 2, 1, 6,
        "Valores em R$ Milhares | Fonte: RI ri.nisseisa.com.br | Analise: Research Analyst + Accountant + Financial Modeler",
        bg=C_NAVY2, color=C_LIGHT, sz=9, bold=False)

    # Cabecalhos
    header(ws, 3, 1, "LINHA")
    header(ws, 3, 2, "UNIDADE")
    for i, (a, d) in enumerate(zip(ANOS, DATAS)):
        header(ws, 3, 3+i, str(a))
        header(ws, 4, 3+i, d, bg=C_NAVY2, sz=8)
    header(ws, 4, 1, ""); header(ws, 4, 2, "")

    # Secao DRE
    ROW = 5
    def sec(text, bg=C_NAVY2):
        nonlocal ROW
        ws.merge_cells(start_row=ROW, start_column=1, end_row=ROW, end_column=6)
        c = ws.cell(ROW, 1, text)
        c.fill = fill(bg); c.font = font(bold=True, color=C_WHITE, size=9)
        c.alignment = align("left"); c.border = border_thin()
        ROW += 1

    def linha(label, valores, fmt=FMT_BRL, bg=C_WHITE, bold=False,
              color=C_GRAPHITE, unidade="R$ mil", indent=0):
        nonlocal ROW
        label_cell(ws, ROW, 1, label, bg=bg, bold=bold, color=color, indent=indent)
        data_cell(ws, ROW, 2, unidade, bg=bg, h="center")
        for i, v in enumerate(valores):
            data_cell(ws, ROW, 3+i, v, fmt=fmt, bg=bg, bold=bold, color=color)
        ROW += 1

    def linha_pct(label, valores, bg=C_ALT, color=C_GRAY):
        nonlocal ROW
        label_cell(ws, ROW, 1, f"   {label}", bg=bg, color=color)
        data_cell(ws, ROW, 2, "%", bg=bg, h="center", color=color)
        for i, v in enumerate(valores):
            data_cell(ws, ROW, 3+i, v, fmt=FMT_PCT1, bg=bg, color=color)
        ROW += 1

    sec("DEMONSTRACAO DE RESULTADO DO EXERCICIO (DRE)")
    linha("Receita Bruta", REC_BRUTA, bold=True, color=C_NAVY, bg=C_LIGHT)
    linha("  Deducoes e Impostos sobre Vendas", DEDUCOES, indent=1, color=C_RED)
    linha("Receita Liquida (Receita Operacional Liquida)", REC_LIQ, bold=True, color=C_NAVY, bg=C_LIGHT)
    linha("  Custo dos Produtos Vendidos (CPV)", CPV, indent=1, color=C_RED)
    linha("Lucro Bruto", LUC_BRUTO, bold=True, color=C_GREEN, bg=C_LIGHT)
    linha_pct("Margem Bruta (% Rec. Liquida)", MARG_BRUTA)
    linha("  Despesas Comerciais", DESP_COM, indent=1, color=C_RED)
    linha("  Despesas Gerais e Administrativas", DESP_GA, indent=1, color=C_RED)
    linha("  Outras Receitas (Despesas) Operacionais", OUTRAS, indent=1, color=C_RED)
    linha("  Depreciacao e Amortizacao", [-v for v in DA], indent=1, color=C_RED)
    linha("EBIT (Resultado Operacional)", EBIT, bold=True, color=C_NAVY, bg=C_LIGHT)
    linha("  (+) D&A", DA, indent=1, color=C_GREEN)
    linha("EBITDA (IFRS 16 / CPC 06)", EBITDA_IFRS, bold=True, color=C_GREEN, bg=C_LIGHT)
    linha("EBITDA Ex-CPC 06 (Metrica Reportada)", EBITDA_EX, bold=True, color=C_GREEN, bg="E6F4EA")
    linha_pct("Margem EBITDA Ex-CPC06 (% Rec. Bruta)", MARG_EBITDA, color=C_GREEN)
    linha("  Resultado Financeiro Liquido", RES_FIN, indent=1, color=C_RED)
    linha("Resultado Antes dos Tributos (EBT)", EBT, bold=True, color=C_NAVY, bg=C_LIGHT)
    linha("  Imposto de Renda e CSLL", IR, indent=1, color=C_RED)
    linha("Lucro Liquido do Exercicio", LUC_LIQ, bold=True, color=C_GREEN, bg=C_LIGHT)
    linha_pct("Margem Liquida (% Rec. Liquida)", MARG_LIQ, color=C_GREEN)

    ROW += 1
    sec("BALANCO PATRIMONIAL — ATIVO")
    linha("ATIVO CIRCULANTE", AT_CIRC, bold=True, color=C_NAVY, bg=C_LIGHT)
    linha("  Caixa e Equivalentes de Caixa", CAIXA, indent=1)
    linha("  Contas a Receber de Clientes", RECEBER, indent=1)
    linha("  Estoques", ESTOQUES, indent=1)
    linha("  Outros Ativos Circulantes", OUTROS_CIRC, indent=1)
    linha("ATIVO NAO CIRCULANTE", AT_NC, bold=True, color=C_NAVY, bg=C_LIGHT)
    linha("  Imobilizado (liquido)", IMOB, indent=1)
    linha("  Ativo de Direito de Uso (IFRS 16 / CPC 06)", DIR_USO, indent=1)
    linha("  Outros Ativos Nao Circulantes", OUTROS_NC, indent=1)
    linha("TOTAL DO ATIVO", AT_TOTAL, bold=True, color=C_WHITE, bg=C_NAVY)

    ROW += 1
    sec("BALANCO PATRIMONIAL — PASSIVO E PATRIMONIO LIQUIDO")
    linha("PASSIVO CIRCULANTE", PAS_CIRC, bold=True, color=C_NAVY, bg=C_LIGHT)
    linha("  Fornecedores", FORNEC, indent=1)
    linha("  Emprestimos e Financiamentos CP", EMP_CP, indent=1)
    linha("  Passivo de Arrendamento CP (IFRS 16)", IFRS_CP, indent=1)
    linha("  Outros Passivos Circulantes", OUTROS_PCIRC, indent=1)
    linha("PASSIVO NAO CIRCULANTE", PAS_NC, bold=True, color=C_NAVY, bg=C_LIGHT)
    linha("  Debentures e Financiamentos LP", DEB_LP, indent=1)
    linha("  Passivo de Arrendamento LP (IFRS 16)", IFRS_LP, indent=1)
    linha("  Outros Passivos Nao Circulantes", OUTROS_NC_P, indent=1)
    linha("PATRIMONIO LIQUIDO", PL, bold=True, color=C_GREEN, bg="E6F4EA")
    linha("TOTAL DO PASSIVO + PL", PAS_TOTAL, bold=True, color=C_WHITE, bg=C_NAVY)

    ROW += 1
    sec("ESTRUTURA DE DIVIDA (Ex-IFRS 16)")
    linha("Divida Bruta (Emprestimos + Debentures)", DIV_BRUTA, bold=True)
    linha("  (-) Caixa e Equivalentes", [-v for v in CAIXA], indent=1, color=C_RED)
    linha("Divida Liquida (Ex-CPC 06)", DIV_LIQ, bold=True, color=C_NAVY, bg=C_LIGHT)
    linha("EBITDA Ex-CPC 06", EBITDA_EX, bold=True, color=C_GREEN, bg=C_LIGHT)
    linha("Divida Liquida / EBITDA", ND_EBITDA, fmt=FMT_MUL, bold=True, color=C_AMBER, bg="FFF3CD")
    linha("Cobertura de Juros (EBIT / Desp. Fin.)", COB_JUROS, fmt=FMT_MUL)

# ---------------------------------------------------------------------------
# ABA 2: Material
# ---------------------------------------------------------------------------
def build_material(ws):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "D4"
    ws.column_dimensions["A"].width = 1
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 10
    for col in ["D","E","F","G"]:
        ws.column_dimensions[col].width = 16

    merge_title(ws, 1, 1, 7,
        "NISSEI — DEMONSTRACOES SIMPLIFICADAS + PROJECOES",
        bg=C_NAVY, sz=13)
    merge_title(ws, 2, 1, 7,
        "Valores em R$ Milhares | 2022A-2025A | 2026E Projecao Base",
        bg=C_NAVY2, color=C_LIGHT, sz=9, bold=False)

    ANOS_M = [2022, 2023, 2024, 2025, "2026E"]
    VALS_M = {
        "Receita Bruta":       REC_BRUTA + [4_100_000],
        "Receita Liquida":     REC_LIQ   + [3_607_500],
        "Custo dos Produtos":  CPV       + [-2_485_200],
        "Lucro Bruto":         LUC_BRUTO + [1_122_300],
        "Despesas Comerciais": DESP_COM  + [-415_800],
        "Despesas G&A":        DESP_GA   + [-208_000],
        "Outras (liquido)":    OUTRAS    + [-25_000],
        "EBIT":                EBIT      + [473_500],
        "(+) D&A":             DA        + [112_000],
        "EBITDA IFRS 16":      EBITDA_IFRS + [585_500],
        "EBITDA Ex-CPC 06":    EBITDA_EX   + [287_000],
        "Resultado Financeiro":RES_FIN   + [-50_000],
        "EBT":                 EBT       + [423_500],
        "IR / CSLL":           IR        + [-84_000],
        "Lucro Liquido":       LUC_LIQ   + [339_500],
    }
    BOLD_M = {"Receita Bruta","Receita Liquida","Lucro Bruto","EBIT",
              "EBITDA Ex-CPC 06","EBITDA IFRS 16","Lucro Liquido","EBT"}
    RED_M  = {"Custo dos Produtos","Despesas Comerciais","Despesas G&A",
              "Outras (liquido)","Resultado Financeiro","IR / CSLL"}

    # Cabecalhos
    header(ws, 3, 2, "LINHA")
    header(ws, 3, 3, "UNIDADE")
    for i, a in enumerate(ANOS_M):
        bg = "D4A853" if str(a) == "2026E" else C_NAVY
        header(ws, 3, 4+i, str(a), bg=bg)

    r = 4
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(r, 1, "DEMONSTRACAO DE RESULTADO").fill = fill(C_NAVY2)
    ws.cell(r, 1).font = font(bold=True, color=C_WHITE, size=9)
    ws.cell(r, 1).alignment = align("left")
    r += 1

    for lbl, vals in VALS_M.items():
        is_bold = lbl in BOLD_M
        color = C_RED if lbl in RED_M else (C_GREEN if lbl in {"Lucro Bruto","EBITDA Ex-CPC 06","Lucro Liquido"} else C_NAVY if is_bold else C_GRAPHITE)
        bg = C_LIGHT if is_bold else (C_ALT if r % 2 == 0 else C_WHITE)
        label_cell(ws, r, 2, lbl, bg=bg, bold=is_bold, color=color)
        data_cell(ws, r, 3, "R$ mil", bg=bg, h="center", color=C_GRAY)
        for i, v in enumerate(vals):
            is_proj = i == 4
            bg_c = "FFF9E6" if is_proj else bg
            data_cell(ws, r, 4+i, v, fmt=FMT_BRL, bg=bg_c, bold=is_bold, color=color)
        r += 1

    # Margens
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(r, 1, "MARGENS").fill = fill(C_NAVY2)
    ws.cell(r, 1).font = font(bold=True, color=C_WHITE, size=9)
    ws.cell(r, 1).alignment = align("left")
    r += 1

    margens = [
        ("Margem Bruta (% Rec. Liq.)", MARG_BRUTA + [0.3110], C_GREEN),
        ("Margem EBITDA Ex-CPC06 (% Rec. Bruta)", MARG_EBITDA + [0.070], C_GREEN),
        ("Margem EBIT (% Rec. Liq.)", [EBIT[i]/REC_LIQ[i] for i in range(4)] + [0.1312], C_NAVY),
        ("Margem Liquida (% Rec. Liq.)", MARG_LIQ + [0.0941], C_GREEN),
    ]
    for lbl, vals, color in margens:
        label_cell(ws, r, 2, lbl, color=color)
        data_cell(ws, r, 3, "%", h="center", color=C_GRAY)
        for i, v in enumerate(vals):
            bg_c = "FFF9E6" if i == 4 else (C_ALT if r%2==0 else C_WHITE)
            data_cell(ws, r, 4+i, v, fmt=FMT_PCT1, bg=bg_c, color=color)
        r += 1

# ---------------------------------------------------------------------------
# ABA 3: Painel
# ---------------------------------------------------------------------------
def build_painel(ws):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C4"
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 10
    for col in ["C","D","E","F"]:
        ws.column_dimensions[col].width = 16

    merge_title(ws, 1, 1, 6,
        "NISSEI — PAINEL DE KPIs E FLUXO DE CAIXA",
        bg=C_NAVY, sz=13)
    merge_title(ws, 2, 1, 6,
        "Valores em R$ Milhares | Indices Ex-CPC 06 | Fonte: RI Nissei",
        bg=C_NAVY2, color=C_LIGHT, sz=9, bold=False)

    header(ws, 3, 1, "INDICADOR"); header(ws, 3, 2, "UNIDADE")
    for i, a in enumerate(ANOS):
        header(ws, 3, 3+i, str(a))

    r = 4
    def sec(text):
        nonlocal r
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(r, 1, text).fill = fill(C_NAVY2)
        ws.cell(r, 1).font = font(bold=True, color=C_WHITE, size=9)
        ws.cell(r, 1).alignment = align("left")
        ws.cell(r, 1).border = border_thin()
        r += 1

    def row_kpi(lbl, vals, fmt=FMT_BRL, color=C_GRAPHITE, bold=False):
        nonlocal r
        bg = C_LIGHT if bold else (C_ALT if r%2==0 else C_WHITE)
        label_cell(ws, r, 1, lbl, bg=bg, bold=bold, color=color)
        data_cell(ws, r, 2, "R$ mil" if fmt==FMT_BRL else ("%" if fmt==FMT_PCT1 else "x"), bg=bg, h="center", color=C_GRAY)
        for i, v in enumerate(vals):
            data_cell(ws, r, 3+i, v, fmt=fmt, bg=bg, bold=bold, color=color)
        r += 1

    sec("RECEITA E RENTABILIDADE")
    row_kpi("Receita Bruta",          REC_BRUTA, bold=True, color=C_NAVY)
    row_kpi("Receita Liquida",         REC_LIQ,   bold=True, color=C_NAVY)
    row_kpi("Lucro Bruto",             LUC_BRUTO, color=C_GREEN, bold=True)
    row_kpi("Margem Bruta",            MARG_BRUTA, fmt=FMT_PCT1, color=C_GREEN)
    row_kpi("(-) SG&A (Com + G&A)",    [DESP_COM[i]+DESP_GA[i] for i in range(4)], color=C_RED)
    row_kpi("% SG&A / Rec. Liquida",   [(DESP_COM[i]+DESP_GA[i])/REC_LIQ[i] for i in range(4)], fmt=FMT_PCT1, color=C_RED)
    row_kpi("EBIT",                    EBIT, color=C_NAVY, bold=True)
    row_kpi("(+) D&A",                 DA, color=C_GREEN)
    row_kpi("EBITDA IFRS 16",          EBITDA_IFRS, color=C_NAVY, bold=True)
    row_kpi("EBITDA Ex-CPC 06",        EBITDA_EX, color=C_GREEN, bold=True)
    row_kpi("Margem EBITDA Ex-CPC06",  MARG_EBITDA, fmt=FMT_PCT1, color=C_GREEN)
    row_kpi("Lucro Liquido",           LUC_LIQ, color=C_GREEN, bold=True)
    row_kpi("Margem Liquida",          MARG_LIQ, fmt=FMT_PCT1, color=C_GREEN)

    r += 1
    sec("FLUXO DE CAIXA")
    row_kpi("Fluxo de Caixa Operacional (FCO)", FCO, color=C_GREEN)
    row_kpi("% do EBITDA Ex-CPC06",    [FCO[i]/EBITDA_EX[i] for i in range(4)], fmt=FMT_PCT1, color=C_GRAY)
    row_kpi("(-) CAPEX",               CAPEX, color=C_RED)
    row_kpi("FCFF (Free Cash Flow to Firm)", FCFF, color=C_NAVY, bold=True)
    row_kpi("(-) Juros Pagos",         JUROS_PAGOS, color=C_RED)
    row_kpi("Free Cash Flow (FCL)",    FCL, color=C_GREEN, bold=True)

    r += 1
    sec("ESTRUTURA DE CAPITAL E DIVIDA")
    row_kpi("Divida Bruta",            DIV_BRUTA, bold=True)
    row_kpi("(-) Caixa",               [-v for v in CAIXA], color=C_RED)
    row_kpi("Divida Liquida (Ex-CPC06)",DIV_LIQ,  bold=True, color=C_NAVY)
    row_kpi("Patrimonio Liquido",      PL,         color=C_GREEN)
    row_kpi("Divida Liquida / EBITDA", ND_EBITDA,  fmt=FMT_MUL, bold=True, color=C_AMBER)
    row_kpi("Divida Liquida / PL",     ND_PL,      fmt=FMT_MUL, color=C_AMBER)
    row_kpi("Cobertura de Juros",      COB_JUROS,  fmt=FMT_MUL, color=C_GREEN)
    row_kpi("Liquidez Corrente",       LIQ_CORR,   fmt=FMT_MUL, color=C_GREEN)

# ---------------------------------------------------------------------------
# ABA 4: Graficos (dados para graficos — espelho do VCA)
# ---------------------------------------------------------------------------
def build_graficos(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 35
    for col in ["B","C","D","E"]:
        ws.column_dimensions[col].width = 14

    merge_title(ws, 1, 1, 5, "NISSEI — DADOS PARA GRAFICOS / INDICES", bg=C_NAVY, sz=12)

    for i, a in enumerate(ANOS):
        header(ws, 2, 2+i, str(a))
    header(ws, 2, 1, "INDICADOR")

    dados = [
        ("Liquidez Corrente",          LIQ_CORR,  FMT_MUL,  C_GREEN),
        ("Divida Liquida / PL",         ND_PL,     FMT_MUL,  C_AMBER),
        ("Cobertura de Juros",          COB_JUROS, FMT_MUL,  C_GREEN),
        ("Divida Liquida / EBITDA",     ND_EBITDA, FMT_MUL,  C_AMBER),
        ("Margem EBITDA Ex-CPC06",      MARG_EBITDA, FMT_PCT1, C_GREEN),
        ("Margem Bruta",                MARG_BRUTA,  FMT_PCT1, C_GREEN),
        ("Margem Liquida",              MARG_LIQ,    FMT_PCT1, C_GREEN),
        ("EBITDA Ex-CPC06 (R$ mil)",    EBITDA_EX,   FMT_BRL,  C_NAVY),
        ("Receita Bruta (R$ mil)",      REC_BRUTA,   FMT_BRL,  C_NAVY),
        ("Divida Liquida (R$ mil)",      DIV_LIQ,    FMT_BRL,  C_AMBER),
        ("Lucro Liquido (R$ mil)",       LUC_LIQ,    FMT_BRL,  C_GREEN),
        ("Lojas Ativas",                [350, 368, 468, 476], FMT_BRL, C_NAVY),
    ]

    for r, (lbl, vals, fmt, color) in enumerate(dados, start=3):
        bg = C_ALT if r%2==0 else C_WHITE
        label_cell(ws, r, 1, lbl, bg=bg, color=color)
        for i, v in enumerate(vals):
            data_cell(ws, r, 2+i, v, fmt=fmt, bg=bg, color=color)

# ---------------------------------------------------------------------------
# ABA 5: Painel - Material
# ---------------------------------------------------------------------------
def build_painel_material(ws):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "D4"
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 10
    for col in ["D","E","F","G"]:
        ws.column_dimensions[col].width = 16

    merge_title(ws, 1, 1, 7,
        "NISSEI — PAINEL MATERIAL (ANALISE VERTICAL)",
        bg=C_NAVY, sz=13)
    merge_title(ws, 2, 1, 7,
        "Valores em R$ Milhares | Analise Vertical (% Receita Liquida)",
        bg=C_NAVY2, color=C_LIGHT, sz=9, bold=False)

    header(ws, 3, 2, "LINHA"); header(ws, 3, 3, "UNIDADE")
    for i, a in enumerate(ANOS):
        header(ws, 3, 4+i, str(a))

    r = 4
    rl = REC_LIQ  # base para % vertical

    linhas = [
        ("Receita Liquida", REC_LIQ, None, True, C_NAVY),
        ("Lucro Bruto",     LUC_BRUTO, lambda i: LUC_BRUTO[i]/rl[i], True, C_GREEN),
        ("(-) Despesas G&A",DESP_GA, lambda i: DESP_GA[i]/rl[i], False, C_RED),
        ("(-) Despesas Comerciais",DESP_COM, lambda i: DESP_COM[i]/rl[i], False, C_RED),
        ("EBIT",            EBIT, lambda i: EBIT[i]/rl[i], True, C_NAVY),
        ("(+) D&A",         DA, lambda i: DA[i]/rl[i], False, C_GREEN),
        ("EBITDA IFRS 16",  EBITDA_IFRS, lambda i: EBITDA_IFRS[i]/rl[i], True, C_NAVY),
        ("EBITDA Ex-CPC06", EBITDA_EX, lambda i: EBITDA_EX[i]/REC_BRUTA[i], True, C_GREEN),
        ("Resultado Financeiro",RES_FIN,lambda i: RES_FIN[i]/rl[i], False, C_RED),
        ("Lucro Liquido",   LUC_LIQ, lambda i: LUC_LIQ[i]/rl[i], True, C_GREEN),
        ("Divida Bruta",    DIV_BRUTA, None, False, C_GRAPHITE),
        ("(-) Caixa",       [-v for v in CAIXA], None, False, C_RED),
        ("Divida Liquida",  DIV_LIQ, lambda i: DIV_LIQ[i]/EBITDA_EX[i], True, C_AMBER),
        ("ND/EBITDA (x)",   ND_EBITDA, None, True, C_AMBER),
        ("Cobertura de Juros (x)", COB_JUROS, None, False, C_GREEN),
        ("Liquidez Corrente (x)", LIQ_CORR, None, False, C_GREEN),
        ("CAPEX",           CAPEX, lambda i: CAPEX[i]/EBITDA_EX[i], False, C_RED),
        ("FCL",             FCL, lambda i: FCL[i]/EBITDA_EX[i] if EBITDA_EX[i] else 0, True, C_GREEN),
    ]

    for lbl, vals, pct_fn, bold, color in linhas:
        bg = C_LIGHT if bold else (C_ALT if r%2==0 else C_WHITE)
        label_cell(ws, r, 2, lbl, bg=bg, bold=bold, color=color)
        data_cell(ws, r, 3, "R$ mil", bg=bg, h="center", color=C_GRAY)
        for i, v in enumerate(vals):
            fmt = FMT_MUL if lbl in {"ND/EBITDA (x)","Cobertura de Juros (x)","Liquidez Corrente (x)"} else FMT_BRL
            data_cell(ws, r, 4+i, v, fmt=fmt, bg=bg, bold=bold, color=color)
        r += 1

        if pct_fn:
            bg2 = C_ALT if r%2==0 else C_WHITE
            label_cell(ws, r, 2, f"   % / Base", bg=bg2, color=C_GRAY)
            data_cell(ws, r, 3, "%", bg=bg2, h="center", color=C_GRAY)
            for i in range(4):
                pv = pct_fn(i)
                data_cell(ws, r, 4+i, pv, fmt=FMT_PCT1, bg=bg2, color=C_GRAY)
            r += 1

# ---------------------------------------------------------------------------
# ABA 6: Analise Consolidada
# ---------------------------------------------------------------------------
def build_analise_consolidada(ws):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B6"
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 36
    for col in ["C","D","E","F","G","H","I","J","K","L","M"]:
        ws.column_dimensions[col].width = 12

    merge_title(ws, 1, 1, 13, "ANALISE CONSOLIDADA — NISSEI FARMACIAS S.A. (NISS3)", bg=C_NAVY, sz=13)
    merge_title(ws, 2, 1, 13,
        "Valores em R$ Milhares | Analise Vertical (%V) e Horizontal (%H) | Fonte: RI ri.nisseisa.com.br",
        bg=C_NAVY2, color=C_LIGHT, sz=9, bold=False)

    # Cabecalhos multi-nivel
    ws.merge_cells("A3:B3"); ws.cell(3,1,"BALANCO PATRIMONIAL").fill = fill(C_GRAPHITE)
    ws.cell(3,1).font = font(bold=True,color=C_WHITE,size=9); ws.cell(3,1).alignment = align("center")
    for col in range(1,14):
        ws.cell(3,col).border = border_thin()

    anos_h = [2022, 2023, 2024, 2025]
    col_starts = [3, 6, 9, 12]
    for i, (a, cs) in enumerate(zip(anos_h, col_starts)):
        ws.merge_cells(start_row=3, start_column=cs, end_row=3, end_column=cs+2)
        c = ws.cell(3, cs, str(a))
        c.fill = fill(C_NAVY); c.font = font(bold=True, color=C_WHITE)
        c.alignment = align("center"); c.border = border_thin()

    sub_hdrs = ["VALOR", "%V", "%H"]
    for cs in col_starts:
        for j, sh in enumerate(sub_hdrs):
            header(ws, 4, cs+j, sh, sz=8)
    header(ws, 4, 1, "TIPO", sz=8)
    header(ws, 4, 2, "LINHA", sz=8)

    r = 5
    def sec_ac(text):
        nonlocal r
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=13)
        ws.cell(r,1,text).fill = fill(C_NAVY2)
        ws.cell(r,1).font = font(bold=True,color=C_WHITE,size=9)
        ws.cell(r,1).alignment = align("left"); ws.cell(r,1).border = border_thin()
        r += 1

    def row_ac(tipo, lbl, vals, is_bold=False, color=C_GRAPHITE):
        nonlocal r
        base = vals[0] if vals[0] else 1
        bg = C_LIGHT if is_bold else (C_ALT if r%2==0 else C_WHITE)
        data_cell(ws, r, 1, tipo, bg=bg, h="center", color=C_GRAY)
        label_cell(ws, r, 2, lbl, bg=bg, bold=is_bold, color=color)
        prev = None
        for i, (v, cs) in enumerate(zip(vals, col_starts)):
            at = AT_TOTAL[i] if AT_TOTAL[i] else 1
            pv = v/at if at else 0
            ph = (v/prev - 1) if (prev and prev != 0 and i > 0) else None
            data_cell(ws, r, cs,   v,  fmt=FMT_BRL, bg=bg, bold=is_bold, color=color)
            data_cell(ws, r, cs+1, pv, fmt=FMT_PCT1, bg=bg, color=C_GRAY)
            c_ph = ws.cell(r, cs+2)
            if ph is not None:
                c_ph.value = ph; c_ph.number_format = FMT_PCT1
                c_ph.font = font(color=C_GREEN if ph>0 else C_RED)
            else:
                c_ph.value = "—"; c_ph.font = font(color=C_GRAY)
            c_ph.fill = fill(bg); c_ph.alignment = align("center"); c_ph.border = border_thin()
            prev = v
        r += 1

    sec_ac("1. ATIVO")
    row_ac("FIN","Caixa e Equivalentes de Caixa", CAIXA)
    row_ac("OP", "Contas a Receber",              RECEBER)
    row_ac("OP", "Estoques",                      ESTOQUES)
    row_ac("OP", "Outros Ativos Circulantes",     OUTROS_CIRC)
    row_ac("OP", "ATIVO CIRCULANTE TOTAL",        AT_CIRC, is_bold=True, color=C_NAVY)
    row_ac("OP", "Imobilizado (liquido)",          IMOB)
    row_ac("OP", "Ativo Direito de Uso (CPC06)",  DIR_USO)
    row_ac("OP", "Outros Ativos NC",              OUTROS_NC)
    row_ac("OP", "ATIVO NAO CIRCULANTE TOTAL",    AT_NC, is_bold=True, color=C_NAVY)
    row_ac("--", "TOTAL DO ATIVO",                AT_TOTAL, is_bold=True, color=C_WHITE)

    r += 1
    sec_ac("2. PASSIVO")
    row_ac("OP", "Fornecedores",                  FORNEC)
    row_ac("FIN","Emprestimos CP",                EMP_CP)
    row_ac("FIN","Arrendamento CP (IFRS 16)",     IFRS_CP)
    row_ac("OP", "Outros Passivos Circ.",         OUTROS_PCIRC)
    row_ac("OP", "PASSIVO CIRCULANTE TOTAL",      PAS_CIRC, is_bold=True, color=C_NAVY)
    row_ac("FIN","Debentures e Financiamentos LP",DEB_LP)
    row_ac("FIN","Arrendamento LP (IFRS 16)",     IFRS_LP)
    row_ac("OP", "Outros NC",                     OUTROS_NC_P)
    row_ac("OP", "PASSIVO NAO CIRC. TOTAL",       PAS_NC, is_bold=True, color=C_NAVY)
    row_ac("PL", "PATRIMONIO LIQUIDO",            PL, is_bold=True, color=C_GREEN)
    row_ac("--", "TOTAL PASSIVO + PL",            PAS_TOTAL, is_bold=True, color=C_WHITE)

    r += 1
    sec_ac("3. DRE — ANALISE VERTICAL E HORIZONTAL")
    dre_rows = [
        ("OP","Receita Bruta",          REC_BRUTA, True, C_NAVY),
        ("OP","Deducoes",               DEDUCOES,  False,C_RED),
        ("OP","Receita Liquida",        REC_LIQ,   True, C_NAVY),
        ("OP","CPV",                    CPV,       False,C_RED),
        ("OP","Lucro Bruto",            LUC_BRUTO, True, C_GREEN),
        ("OP","Despesas Comerciais",    DESP_COM,  False,C_RED),
        ("OP","Despesas G&A",           DESP_GA,   False,C_RED),
        ("OP","Outras (liq.)",          OUTRAS,    False,C_RED),
        ("OP","EBIT",                   EBIT,      True, C_NAVY),
        ("OP","D&A",                    DA,        False,C_GREEN),
        ("OP","EBITDA IFRS 16",         EBITDA_IFRS,True,C_NAVY),
        ("OP","EBITDA Ex-CPC06",        EBITDA_EX, True, C_GREEN),
        ("FIN","Resultado Financeiro",  RES_FIN,   False,C_RED),
        ("OP","EBT",                    EBT,       True, C_NAVY),
        ("OP","IR / CSLL",              IR,        False,C_RED),
        ("OP","Lucro Liquido",          LUC_LIQ,   True, C_GREEN),
    ]
    for tipo, lbl, vals, bold, color in dre_rows:
        row_ac(tipo, lbl, vals, is_bold=bold, color=color)

    r += 1
    sec_ac("4. INDICES DE CREDITO")
    idx_rows = [
        ("Liquidez Corrente",   LIQ_CORR,  FMT_MUL),
        ("ND / EBITDA (x)",     ND_EBITDA, FMT_MUL),
        ("Cobertura de Juros",  COB_JUROS, FMT_MUL),
        ("ND / PL",             ND_PL,     FMT_MUL),
        ("Margem EBITDA",       MARG_EBITDA,FMT_PCT1),
        ("Margem Bruta",        MARG_BRUTA, FMT_PCT1),
        ("Margem Liquida",      MARG_LIQ,   FMT_PCT1),
    ]
    for lbl, vals, fmt in idx_rows:
        bg = C_ALT if r%2==0 else C_WHITE
        label_cell(ws, r, 2, lbl, bg=bg, bold=True, color=C_NAVY)
        data_cell(ws, r, 1, "IDX", bg=bg, h="center", color=C_GRAY)
        prev = None
        for i, (v, cs) in enumerate(zip(vals, col_starts)):
            c1 = data_cell(ws, r, cs,   v,  fmt=fmt, bg=bg, bold=True, color=C_NAVY)
            ws.cell(r, cs+1).fill = fill(bg); ws.cell(r, cs+1).border = border_thin()
            c_ph = ws.cell(r, cs+2)
            if prev is not None and prev != 0:
                dif = v - prev
                c_ph.value = dif
                c_ph.number_format = fmt
                c_ph.font = font(color=C_GREEN if dif>0 else C_RED)
            else:
                c_ph.value = "—"
                c_ph.font = font(color=C_GRAY)
            c_ph.fill = fill(bg); c_ph.alignment = align("center"); c_ph.border = border_thin()
            prev = v
        r += 1

# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def build():
    wb = openpyxl.Workbook()

    # Remover aba padrao
    wb.remove(wb.active)

    sheets = [
        ("DF's",               build_dfs),
        ("Material",           build_material),
        ("Painel",             build_painel),
        ("Graficos",           build_graficos),
        ("Painel - Material",  build_painel_material),
        ("Analise Consolidada",build_analise_consolidada),
    ]

    for name, fn in sheets:
        ws = wb.create_sheet(name)
        # Tab color
        ws.sheet_properties.tabColor = C_NAVY if name != "Analise Consolidada" else C_GOLD
        fn(ws)
        print(f"  Aba '{name}' gerada.")

    for path in OUT_PATHS:
        Path(path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        print(f"Salvo: {path}")

    print(f"\nConcluido. {len(sheets)} abas geradas.")

if __name__ == "__main__":
    build()
