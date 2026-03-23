"""
KnowledgeIngestor
Processa arquivos de referência (PDF, Excel, DOCX, TXT) e os indexa
no ChromaDB para uso dos agentes via RAG.

Estratégia de chunking:
  - PDF: por página + seções detectadas (títulos em maiúsculas, negrito)
  - Excel: por aba, convertida para texto estruturado
  - DOCX: por parágrafo agrupado
  - TXT/MD: por bloco de ~500 tokens

Embeddings:
  - Primário: sentence-transformers (multilingual-e5-base)
  - Fallback: TF-IDF com sklearn (sem GPU, sem API key)
"""

from __future__ import annotations

import hashlib
import json
import os
import re
from pathlib import Path
from typing import Any

from loguru import logger


KNOWLEDGE_DB_PATH = os.getenv(
    "KNOWLEDGE_DB_PATH",
    str(Path(__file__).parent.parent.parent / "knowledge" / "chroma_db"),
)
KNOWLEDGE_COLLECTION = "ib_reference_knowledge"
CHUNK_SIZE = 600       # tokens aproximados por chunk
CHUNK_OVERLAP = 80


class KnowledgeIngestor:
    """Ingests reference documents into the knowledge ChromaDB collection."""

    def __init__(self, db_path: str = KNOWLEDGE_DB_PATH):
        self.db_path = db_path
        self._client = None
        self._collection = None
        self._embedder = None

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def ingest_directory(self, directory: str) -> dict[str, Any]:
        """Walk directory recursively and ingest all supported files."""
        base = Path(directory)
        if not base.exists():
            return {"error": f"Diretório não encontrado: {directory}"}

        stats = {"processed": 0, "chunks_added": 0, "skipped": 0, "errors": []}
        supported = {".pdf", ".xlsx", ".xls", ".docx", ".txt", ".md"}

        for file_path in sorted(base.rglob("*")):
            if file_path.suffix.lower() not in supported:
                continue
            if file_path.name.startswith(".") or file_path.name.startswith("~"):
                continue
            try:
                n = self.ingest_file(str(file_path))
                stats["chunks_added"] += n
                stats["processed"] += 1
                logger.info(f"Indexado: {file_path.name} ({n} chunks)")
            except Exception as exc:
                stats["errors"].append({"file": file_path.name, "error": str(exc)})
                stats["skipped"] += 1
                logger.warning(f"Erro ao indexar {file_path.name}: {exc}")

        return stats

    def ingest_file(self, file_path: str) -> int:
        """Ingest a single file. Returns number of chunks added."""
        path = Path(file_path)
        suffix = path.suffix.lower()

        chunks = []
        if suffix == ".pdf":
            chunks = self._extract_pdf(path)
        elif suffix in (".xlsx", ".xls"):
            chunks = self._extract_excel(path)
        elif suffix == ".docx":
            chunks = self._extract_docx(path)
        elif suffix in (".txt", ".md"):
            chunks = self._extract_text(path)
        else:
            raise ValueError(f"Formato não suportado: {suffix}")

        if not chunks:
            return 0

        collection = self._get_collection()
        category = self._detect_category(path)

        is_modelo = "Modelo" in path.parts

        docs, ids, metas = [], [], []
        for i, chunk in enumerate(chunks):
            chunk_id = self._make_id(str(path), i)
            docs.append(chunk["text"])
            ids.append(chunk_id)
            metas.append({
                "source_file": path.name,
                "source_path": str(path),
                "category": category,
                "chunk_index": i,
                "total_chunks": len(chunks),
                "page": chunk.get("page", ""),
                "section": chunk.get("section", ""),
                "is_modelo": "true" if is_modelo else "false",
            })

        # Upsert in batches of 50
        batch = 50
        for i in range(0, len(docs), batch):
            collection.upsert(
                documents=docs[i:i+batch],
                ids=ids[i:i+batch],
                metadatas=metas[i:i+batch],
            )

        return len(chunks)

    def list_indexed(self) -> list[dict]:
        """Return summary of indexed documents."""
        try:
            col = self._get_collection()
            result = col.get(include=["metadatas"])
            seen: dict[str, dict] = {}
            for meta in result.get("metadatas", []):
                fname = meta.get("source_file", "")
                if fname not in seen:
                    seen[fname] = {
                        "file": fname,
                        "category": meta.get("category", ""),
                        "is_modelo": meta.get("is_modelo", "false"),
                        "chunks": 0,
                    }
                seen[fname]["chunks"] += 1
            # Arquivos Modelo aparecem primeiro na listagem
            return sorted(seen.values(), key=lambda x: (x["is_modelo"] != "true", x["file"]))
        except Exception as exc:
            return [{"error": str(exc)}]

    def reset(self) -> bool:
        """Delete and recreate the knowledge collection."""
        try:
            client = self._get_client()
            try:
                client.delete_collection(KNOWLEDGE_COLLECTION)
            except Exception:
                pass
            self._collection = None
            self._get_collection()
            logger.info("Knowledge collection resetada.")
            return True
        except Exception as exc:
            logger.error(f"Erro ao resetar: {exc}")
            return False

    def search(self, query: str, n_results: int = 5, category: str = "") -> list[dict]:
        """Search the knowledge base. Results from pasta Modelo are boosted to the top."""
        col = self._get_collection()
        where = {"category": category} if category else None
        # Busca o dobro para garantir que arquivos Modelo apareçam mesmo com score menor
        fetch = max(n_results * 2, 10)
        kwargs: dict = {"query_texts": [query], "n_results": fetch}
        if where:
            kwargs["where"] = where

        res = col.query(**kwargs)
        docs = res.get("documents", [[]])[0]
        metas = res.get("metadatas", [[]])[0]
        dists = res.get("distances", [[]])[0]

        results = []
        for doc, meta, dist in zip(docs, metas, dists):
            is_modelo = meta.get("is_modelo", "false") == "true"
            relevance = round(1 - dist, 3)
            # Boost de 0.15 para arquivos da pasta Modelo
            boosted = min(round(relevance + 0.15, 3), 1.0) if is_modelo else relevance
            results.append({
                "text": doc,
                "source": meta.get("source_file", ""),
                "category": meta.get("category", ""),
                "page": meta.get("page", ""),
                "section": meta.get("section", ""),
                "relevance": boosted,
                "is_modelo": is_modelo,
            })

        # Ordenar por relevance boosted, Modelo primeiro em caso de empate
        results.sort(key=lambda x: (-x["relevance"], not x["is_modelo"]))
        return results[:n_results]

    # ------------------------------------------------------------------
    # Extractors
    # ------------------------------------------------------------------

    def _extract_pdf(self, path: Path) -> list[dict]:
        chunks = []
        try:
            import pdfplumber
            with pdfplumber.open(str(path)) as pdf:
                for page_num, page in enumerate(pdf.pages, 1):
                    text = page.extract_text() or ""
                    # Also extract tables as structured text
                    for table in page.extract_tables():
                        text += "\n" + self._table_to_text(table)
                    if text.strip():
                        for chunk in self._split_text(text, CHUNK_SIZE, CHUNK_OVERLAP):
                            chunks.append({
                                "text": chunk,
                                "page": str(page_num),
                                "section": self._detect_section(chunk),
                            })
        except ImportError:
            # Fallback: try pypdf
            chunks = self._extract_pdf_fallback(path)
        return chunks

    def _extract_pdf_fallback(self, path: Path) -> list[dict]:
        try:
            import pypdf
            chunks = []
            with open(str(path), "rb") as f:
                reader = pypdf.PdfReader(f)
                for i, page in enumerate(reader.pages):
                    text = page.extract_text() or ""
                    for chunk in self._split_text(text, CHUNK_SIZE, CHUNK_OVERLAP):
                        chunks.append({"text": chunk, "page": str(i+1), "section": ""})
            return chunks
        except ImportError:
            logger.warning(f"Nem pdfplumber nem pypdf disponíveis para {path.name}. Instale: pip install pdfplumber")
            return []

    def _extract_excel(self, path: Path) -> list[dict]:
        import openpyxl
        chunks = []
        try:
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows_text = []
                for row in ws.iter_rows(max_row=200, values_only=True):
                    row_vals = [str(v) if v is not None else "" for v in row]
                    non_empty = [v for v in row_vals if v.strip()]
                    if non_empty:
                        rows_text.append(" | ".join(row_vals))
                sheet_text = f"[Aba: {sheet_name}]\n" + "\n".join(rows_text)
                for chunk in self._split_text(sheet_text, CHUNK_SIZE, CHUNK_OVERLAP):
                    chunks.append({"text": chunk, "page": sheet_name, "section": sheet_name})
            wb.close()
        except Exception as exc:
            logger.warning(f"Erro ao ler Excel {path.name}: {exc}")
        return chunks

    def _extract_docx(self, path: Path) -> list[dict]:
        try:
            import docx
            doc = docx.Document(str(path))
            paragraphs = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
            full_text = "\n".join(paragraphs)
            return [
                {"text": c, "page": "", "section": self._detect_section(c)}
                for c in self._split_text(full_text, CHUNK_SIZE, CHUNK_OVERLAP)
            ]
        except ImportError:
            logger.warning(f"python-docx não instalado. Instale: pip install python-docx")
            return []

    def _extract_text(self, path: Path) -> list[dict]:
        text = path.read_text(encoding="utf-8", errors="ignore")
        return [
            {"text": c, "page": "", "section": self._detect_section(c)}
            for c in self._split_text(text, CHUNK_SIZE, CHUNK_OVERLAP)
        ]

    # ------------------------------------------------------------------
    # Chunking
    # ------------------------------------------------------------------

    def _split_text(self, text: str, chunk_size: int, overlap: int) -> list[str]:
        """Split text into overlapping chunks by word count approximation."""
        words = text.split()
        if not words:
            return []
        chunks = []
        start = 0
        while start < len(words):
            end = min(start + chunk_size, len(words))
            chunk = " ".join(words[start:end]).strip()
            if len(chunk) > 50:  # skip tiny chunks
                chunks.append(chunk)
            start += chunk_size - overlap
        return chunks

    def _table_to_text(self, table: list) -> str:
        if not table:
            return ""
        rows = []
        for row in table:
            if row:
                rows.append(" | ".join(str(c or "") for c in row))
        return "\n".join(rows)

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _detect_category(self, path: Path) -> str:
        parts = path.parts
        for part in parts:
            part_lower = part.lower()
            if "pitch" in part_lower:
                return "pitch_book"
            if "financial_model" in part_lower or "modelo" in part_lower:
                return "financial_model"
            if "research" in part_lower or "report" in part_lower:
                return "research_report"
        # Guess from file extension
        if path.suffix.lower() in (".xlsx", ".xls"):
            return "financial_model"
        return "general"

    def _detect_section(self, text: str) -> str:
        """Try to detect the section name from text content."""
        section_keywords = [
            "executive summary", "sumário executivo", "valuation", "valuation summary",
            "financial highlights", "destaques financeiros", "investment thesis",
            "tese de investimento", "risk", "riscos", "appendix", "apêndice",
            "transaction overview", "company overview", "visão geral",
            "ebitda", "dcf", "lbo", "comps", "comparable",
        ]
        text_lower = text[:200].lower()
        for kw in section_keywords:
            if kw in text_lower:
                return kw.title()
        return ""

    def _make_id(self, file_path: str, chunk_index: int) -> str:
        h = hashlib.md5(f"{file_path}:{chunk_index}".encode()).hexdigest()[:12]
        return f"doc_{h}_{chunk_index}"

    def _get_client(self):
        if self._client is None:
            import chromadb
            Path(self.db_path).mkdir(parents=True, exist_ok=True)
            self._client = chromadb.PersistentClient(path=self.db_path)
        return self._client

    def _get_collection(self):
        if self._collection is None:
            client = self._get_client()
            # Use default embedding (chromadb built-in) — no API key needed
            self._collection = client.get_or_create_collection(
                name=KNOWLEDGE_COLLECTION,
                metadata={"hnsw:space": "cosine"},
            )
        return self._collection
